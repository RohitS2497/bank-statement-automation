"""
Bank-statement automation pipeline (MVP).

Stages:  fetch -> parse -> analyse -> visualise -> cross-check -> mask(optional) -> edge-cases
Input :  folder of PDFs (what the Apps Script drops into your Drive)
Output:  clean CSV, multi-sheet Excel, charts, and a run log.

Designed to run locally OR in Google Colab (mount Drive, point INBOX at the folder).
"""
import os, re, sys, logging, datetime as dt
from dataclasses import dataclass, field
import pdfplumber
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from dateutil import parser as dateparser
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment

# ----------------------------- config -----------------------------
HERE     = os.path.dirname(os.path.abspath(__file__))
INBOX    = os.path.join(HERE, "inbox")
OUT      = os.path.join(HERE, "outputs")
CHARTS   = os.path.join(OUT, "charts")
MASK_PII = False          # user said dummy data -> off. Flip to True to redact.
PDF_PASSWORDS = []   # add known passwords here to auto-unlock protected PDFs
os.makedirs(CHARTS, exist_ok=True)

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)-7s | %(message)s",
                    handlers=[logging.StreamHandler(sys.stdout),
                              logging.FileHandler(os.path.join(OUT, "run_log.txt"), mode="w")])
log = logging.getLogger("pipeline")

# canonical column -> list of possible header names seen across banks
HEADER_MAP = {
    "date":    ["date", "txn date", "transaction date", "value date", "posting date"],
    "desc":    ["narration", "particulars", "description", "details", "remarks"],
    "debit":   ["withdrawal (dr)", "withdrawal", "debit", "dr", "money out", "paid out"],
    "credit":  ["deposit (cr)", "deposit", "credit", "cr", "money in", "paid in"],
    "balance": ["closing balance", "balance", "running balance", "available balance"],
}
CATEGORY_RULES = [
    ("Salary/Income", r"salary|payroll|stipend"),
    ("Interest/Refund", r"interest|int\.cr|refund|cashback|settlement"),
    ("Rent", r"rent|landlord"),
    ("Loan/EMI", r"emi|loan"),
    ("Insurance", r"insurance|life|policy"),
    ("Utilities", r"electricity|bses|water|gas|airtel|jio|mobile|broadband|billpay"),
    ("Groceries", r"dmart|bigbasket|grocery|reliance fresh|blinkit|zepto"),
    ("Dining", r"swiggy|zomato|restaurant|cafe|dining"),
    ("Transport", r"uber|ola|fuel|indian oil|petrol|irctc|metro"),
    ("Shopping", r"amazon|flipkart|myntra|shopping|pos"),
    ("Subscriptions", r"netflix|prime|spotify|hotstar|subscription"),
    ("Cash/ATM", r"atm|cash withdrawal"),
    ("Transfers", r"upi|neft|imps|rtgs|transfer"),
]

@dataclass
class FileResult:
    name: str
    status: str          # PARSED / EMPTY / ENCRYPTED / NO_TABLE / ERROR
    pages: int = 0
    rows: int = 0
    note: str = ""

# ----------------------------- helpers -----------------------------
def _norm(s): return re.sub(r"\s+", " ", (s or "").strip().lower())

def _to_amount(x):
    if x is None: return 0.0
    x = str(x).replace(",", "").replace("INR", "").replace("₹", "").strip()
    if x in ("", "-"): return 0.0
    neg = x.startswith("(") and x.endswith(")")
    x = x.strip("()")
    try: v = float(x)
    except ValueError: return 0.0
    return -v if neg else v

def _parse_date(x):
    try: return dateparser.parse(str(x), dayfirst=True).date()
    except Exception: return None

def _map_headers(header_row):
    """Return {col_index: canonical_name} or None if this isn't a real header."""
    mapping, hits = {}, 0
    for i, cell in enumerate(header_row):
        c = _norm(cell)
        for canon, names in HEADER_MAP.items():
            if c in names:
                mapping[i] = canon; hits += 1; break
    # need at least date + description + one money column to trust it
    canon_found = set(mapping.values())
    if {"date", "desc"} <= canon_found and ({"debit", "credit", "balance"} & canon_found):
        return mapping
    return None

def categorise(desc):
    d = _norm(desc)
    for cat, pat in CATEGORY_RULES:
        if re.search(pat, d): return cat
    return "Other"

def mask_text(s):
    if not MASK_PII or not s: return s
    s = re.sub(r"\b(\d[\d ]{6,}\d)\b", lambda m: re.sub(r"\d", "X", m.group()), s)  # long digit runs
    s = re.sub(r"([A-Z]{4}0[A-Z0-9]{6})", "XXXX0XXXXXX", s)                          # IFSC
    return s

# ----------------------------- 1) PARSE (+ edge cases) -----------------------------
def parse_pdf(path):
    name = os.path.basename(path)
    if os.path.getsize(path) < 400:
        return [], FileResult(name, "EMPTY", note="file too small / no content")
    try:
        pdf = pdfplumber.open(path)
    except Exception:
        pdf = None
        for pw in PDF_PASSWORDS:                 # try known passwords for encrypted files
            try:
                pdf = pdfplumber.open(path, password=pw); break
            except Exception: pdf = None
        if pdf is None:
            return [], FileResult(name, "ENCRYPTED", note="password-protected; not in known list")
    txns, header, npages, any_text = [], None, 0, False
    try:
        npages = len(pdf.pages)
        for page in pdf.pages:
            if (page.extract_text() or "").strip(): any_text = True
            table = page.extract_table()
            if not table: continue
            for row in table:
                m = _map_headers(row)
                if m:                       # (re)set header; skip repeated headers on new pages
                    header = m; continue
                if header is None: continue
                rec = {"date": None, "desc": "", "debit": 0.0, "credit": 0.0, "balance": None}
                for idx, canon in header.items():
                    val = row[idx] if idx < len(row) else None
                    if canon == "date":    rec["date"] = _parse_date(val)
                    elif canon == "desc":  rec["desc"] = (val or "").strip()
                    elif canon == "debit": rec["debit"] = _to_amount(val)
                    elif canon == "credit":rec["credit"] = _to_amount(val)
                    elif canon == "balance": rec["balance"] = _to_amount(val) if val not in (None,"") else None
                if rec["date"] and rec["desc"]:
                    rec["source_file"] = name
                    txns.append(rec)
    finally:
        pdf.close()
    if not txns:
        if not any_text:
            return [], FileResult(name, "EMPTY", pages=npages,
                                  note="no text layer (blank/scanned) — needs OCR")
        return [], FileResult(name, "NO_TABLE", pages=npages, note="text present but no transaction table")
    return txns, FileResult(name, "PARSED", pages=npages, rows=len(txns))

# ----------------------------- 3) ANALYSE -----------------------------
def analyse(df):
    df = df.copy()
    df["category"] = df["desc"].map(categorise)
    df["net"] = df["credit"] - df["debit"]
    df["month"] = pd.to_datetime(df["date"]).dt.to_period("M").astype(str)
    cat = (df.groupby("category")
             .agg(spend=("debit", "sum"), income=("credit", "sum"), count=("desc", "size"))
             .sort_values("spend", ascending=False).reset_index())
    monthly = (df.groupby("month")
                 .agg(income=("credit", "sum"), expense=("debit", "sum"), net=("net", "sum"),
                      txns=("desc", "size")).reset_index())
    return df, cat, monthly

# ----------------------------- 4) VISUALISE -----------------------------
def visualise(df, cat, monthly):
    plt.rcParams.update({"font.size": 9, "axes.edgecolor": "#C9D2DA"})
    NAVY, GOLD = "#0C1C2C", "#B4924F"
    paths = {}
    # spend by category
    sc = cat[cat["spend"] > 0].sort_values("spend")
    fig, ax = plt.subplots(figsize=(6.6, 3.6))
    ax.barh(sc["category"], sc["spend"], color=NAVY)
    ax.set_title("Spend by category (INR)", color=NAVY, fontweight="bold")
    ax.spines[["top", "right"]].set_visible(False)
    for i, v in enumerate(sc["spend"]): ax.text(v, i, f" {v:,.0f}", va="center", fontsize=8)
    fig.tight_layout(); p = os.path.join(CHARTS, "spend_by_category.png"); fig.savefig(p, dpi=140); plt.close(fig); paths["cat"] = p
    # income vs expense by month
    fig, ax = plt.subplots(figsize=(6.6, 3.2))
    x = range(len(monthly)); w = 0.38
    ax.bar([i - w/2 for i in x], monthly["income"], w, label="Income", color=GOLD)
    ax.bar([i + w/2 for i in x], monthly["expense"], w, label="Expense", color=NAVY)
    ax.set_xticks(list(x)); ax.set_xticklabels(monthly["month"])
    ax.set_title("Income vs expense by month (INR)", color=NAVY, fontweight="bold")
    ax.spines[["top", "right"]].set_visible(False); ax.legend(frameon=False)
    fig.tight_layout(); p = os.path.join(CHARTS, "income_vs_expense.png"); fig.savefig(p, dpi=140); plt.close(fig); paths["ie"] = p
    # balance trend
    d = df.dropna(subset=["balance"]).sort_values("date")
    fig, ax = plt.subplots(figsize=(6.6, 3.2))
    ax.plot(pd.to_datetime(d["date"]), d["balance"], color=NAVY, lw=1.6)
    ax.fill_between(pd.to_datetime(d["date"]), d["balance"], color=GOLD, alpha=0.15)
    ax.set_title("Account balance over time (INR)", color=NAVY, fontweight="bold")
    ax.spines[["top", "right"]].set_visible(False)
    fig.autofmt_xdate(); fig.tight_layout(); p = os.path.join(CHARTS, "balance_trend.png"); fig.savefig(p, dpi=140); plt.close(fig); paths["bal"] = p
    return paths

# ----------------------------- 5) CROSS-CHECK -----------------------------
def cross_check(df):
    flags = []
    for src, g in df.groupby("source_file", sort=False):
        g = g.reset_index(drop=True)
        prev = None
        for i, r in g.iterrows():
            # running-balance reconciliation
            if r["balance"] is not None and prev is not None:
                expected = round(prev - r["debit"] + r["credit"], 2)
                if abs(expected - r["balance"]) > 0.01:
                    flags.append({"file": src, "row": i + 1, "type": "Balance mismatch",
                                  "detail": f"expected {expected:,.2f} but statement shows {r['balance']:,.2f}",
                                  "desc": r["desc"]})
            if r["balance"] is not None: prev = r["balance"]
            # date order
            if i > 0 and r["date"] and g.loc[i-1, "date"] and r["date"] < g.loc[i-1, "date"]:
                flags.append({"file": src, "row": i + 1, "type": "Date out of order",
                              "detail": f"{r['date']} after {g.loc[i-1,'date']}", "desc": r["desc"]})
            # missing money
            if r["debit"] == 0 and r["credit"] == 0:
                flags.append({"file": src, "row": i + 1, "type": "Zero-value txn",
                              "detail": "no debit or credit amount", "desc": r["desc"]})
        # duplicate detection (same date/desc/amount)
        dup = g[g.duplicated(subset=["date", "desc", "debit", "credit"], keep=False)]
        for i, r in dup.iterrows():
            flags.append({"file": src, "row": i + 1, "type": "Possible duplicate",
                          "detail": f"{r['debit'] or r['credit']:,.2f} on {r['date']}", "desc": r["desc"]})
    return pd.DataFrame(flags)

# ----------------------------- 6) EXPORT -----------------------------
def export_excel(df, cat, monthly, flags, filestatus, charts):
    xlsx = os.path.join(OUT, "Bank_Statement_Analysis.xlsx")
    with pd.ExcelWriter(xlsx, engine="openpyxl") as xw:
        pd.DataFrame({"Dashboard": []}).to_excel(xw, sheet_name="Dashboard", index=False)
        out = df.copy()
        out["desc"] = out["desc"].map(mask_text); out["source_file"] = out["source_file"].map(mask_text)
        out = out[["date", "desc", "category", "debit", "credit", "balance", "month", "source_file"]]
        out.columns = ["Date", "Description", "Category", "Debit", "Credit", "Balance", "Month", "Source File"]
        out.to_excel(xw, sheet_name="Transactions", index=False)
        cat.rename(columns=str.title).to_excel(xw, sheet_name="Category Summary", index=False)
        monthly.rename(columns=str.title).to_excel(xw, sheet_name="Monthly Summary", index=False)
        (flags if not flags.empty else pd.DataFrame([{"file": "-", "row": "-",
            "type": "No issues found", "detail": "all checks passed", "desc": "-"}])
         ).rename(columns=str.title).to_excel(xw, sheet_name="Validation Flags", index=False)
        pd.DataFrame([f.__dict__ for f in filestatus]).rename(columns=str.title
            ).to_excel(xw, sheet_name="File Status", index=False)
    # style + embed charts
    wb = load_workbook(xlsx)
    NAVY = "0C1C2C"
    for ws in wb.worksheets:
        if ws.title == "Dashboard": continue
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=NAVY)
            c.alignment = Alignment(horizontal="left", vertical="center")
        ws.freeze_panes = "A2"
        for col in ws.columns:
            w = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, 10), 46)
    dash = wb["Dashboard"]
    dash["A1"] = "Bank Statement Analysis — Dashboard"; dash["A1"].font = Font(bold=True, size=15, color=NAVY)
    dash["A2"] = f"Generated {dt.datetime.now():%Y-%m-%d %H:%M}  ·  SYNTHETIC DEMO DATA"
    dash["A2"].font = Font(italic=True, size=9, color="8F7238")
    anchors = [("cat", "A4"), ("ie", "A24"), ("bal", "A44")]
    for key, anchor in anchors:
        if key in charts: dash.add_image(XLImage(charts[key]), anchor)
    wb.save(xlsx)
    return xlsx

# ----------------------------- MAIN -----------------------------
def run():
    log.info("STAGE 1/6  FETCH  — reading PDFs from %s", INBOX)
    files = sorted(os.path.join(INBOX, f) for f in os.listdir(INBOX) if f.lower().endswith(".pdf"))
    log.info("           found %d candidate file(s)", len(files))

    log.info("STAGE 2/6  PARSE  — extracting transactions (edge cases handled)")
    all_txns, filestatus = [], []
    for path in files:
        txns, res = parse_pdf(path)
        filestatus.append(res)
        icon = {"PARSED": "OK ", "EMPTY": "SKIP", "ENCRYPTED": "SKIP",
                "NO_TABLE": "SKIP", "ERROR": "ERR "}.get(res.status, "?")
        log.info("           [%s] %-32s %-9s %s", icon, res.name, res.status,
                 f"{res.rows} txns / {res.pages}p" if res.status == "PARSED" else res.note)
        all_txns.extend(txns)

    if not all_txns:
        log.error("No transactions parsed from any file. Nothing to analyse."); return
    df = pd.DataFrame(all_txns)
    log.info("           parsed %d transactions from %d statement(s)",
             len(df), sum(1 for f in filestatus if f.status == "PARSED"))

    log.info("STAGE 3/6  ANALYSE — categorising & summarising")
    df, cat, monthly = analyse(df)
    log.info("           income=INR %s  expense=INR %s  net=INR %s",
             f"{df['credit'].sum():,.0f}", f"{df['debit'].sum():,.0f}", f"{df['net'].sum():,.0f}")

    log.info("STAGE 4/6  VISUALISE — building charts")
    charts = visualise(df, cat, monthly)
    for p in charts.values(): log.info("           chart -> %s", os.path.relpath(p, HERE))

    log.info("STAGE 5/6  CROSS-CHECK — reconciling balances & anomalies")
    flags = cross_check(df)
    if flags.empty:
        log.info("           no anomalies found")
    else:
        for _, f in flags.iterrows():
            log.warning("           FLAG  %-18s %s row %s  (%s)", f["type"], f["file"], f["row"], f["detail"])

    log.info("STAGE 6/6  MASK/EXPORT — PII masking=%s; writing outputs", MASK_PII)
    df.to_csv(os.path.join(OUT, "transactions_clean.csv"), index=False)
    xlsx = export_excel(df, cat, monthly, flags, filestatus, charts)
    log.info("           CSV   -> outputs/transactions_clean.csv")
    log.info("           XLSX  -> %s", os.path.relpath(xlsx, HERE))
    log.info("DONE. %d txns · %d flag(s) · %d file(s) skipped as edge cases.",
             len(df), 0 if flags.empty else len(flags),
             sum(1 for f in filestatus if f.status != "PARSED"))

if __name__ == "__main__":
    run()
